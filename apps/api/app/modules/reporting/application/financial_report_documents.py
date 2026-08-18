"""Gera PDF e XLSX usando o mesmo DTO exibido no relatório financeiro."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.modules.commissions.application.queries.financial_report import (
    FinancialReportBeneficiary,
    FinancialReportSummary,
)

SUMMARY_LABELS = {
    "gross_revenue": "Faturamento bruto",
    "receipt_reversals": "Estornos",
    "recognized_revenue": "Faturamento reconhecido",
    "recognized_production": "Produção reconhecida",
    "consultant_commissions": "Comissões de consultores",
    "leader_commissions": "Comissões de lideranças",
    "finalization_commissions": "Comissões de finalização",
    "finalization_leader_commissions": "Líder de finalização",
    "bko_commissions": "Comissões de BKO",
    "total_commissions": "Total de comissões",
    "net_billing": "Faturamento líquido",
    "bonuses": "Bônus",
    "discounts": "Descontos",
    "deferred": "Adiado",
    "paid": "Pago",
    "payable": "A pagar",
}


def financial_report_xlsx(
    summary: FinancialReportSummary, beneficiaries: list[FinancialReportBeneficiary]
) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Resumo"
    overview.append(["Indicador", "Valor"])
    for field, label in SUMMARY_LABELS.items():
        overview.append([label, float(getattr(summary, field))])
    overview.column_dimensions["A"].width = 34
    overview.column_dimensions["B"].width = 20
    for cell in overview[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="44639F")
    for cell in overview["B"][1:]:
        cell.number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"

    detail = workbook.create_sheet("Beneficiários")
    detail.append(
        [
            "Beneficiário",
            "Estratégias",
            "Automática",
            "Manual",
            "Calculada",
            "Acumulado anterior",
            "Bônus",
            "Desconto",
            "Adiado",
            "Pago",
            "A pagar",
            "Status",
        ]
    )
    for item in beneficiaries:
        detail.append(
            [
                item.beneficiary_name,
                ", ".join(item.strategies),
                float(item.automatic_amount),
                float(item.manual_amount),
                float(item.calculated_amount),
                float(item.carryover_amount),
                float(item.bonus_amount),
                float(item.discount_amount),
                float(item.deferred_amount),
                float(item.paid_amount),
                float(item.payable_amount),
                item.status or "NÃO GERADO",
            ]
        )
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="44639F")
        cell.alignment = Alignment(horizontal="center")
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    detail.column_dimensions["A"].width = 28
    detail.column_dimensions["B"].width = 34
    for column in range(3, 12):
        detail.column_dimensions[chr(64 + column)].width = 18
        for row in range(2, detail.max_row + 1):
            detail.cell(row, column).number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def financial_report_pdf(
    summary: FinancialReportSummary, beneficiaries: list[FinancialReportBeneficiary]
) -> bytes:
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Relatório financeiro de comissões",
    )
    styles = getSampleStyleSheet()
    story: list[object] = [
        Paragraph("Relatório financeiro de comissões", styles["Title"]),
        Spacer(1, 5 * mm),
    ]
    summary_table = Table(
        [
            ["Faturamento reconhecido", "Produção", "Comissões", "Faturamento líquido"],
            [
                _money(summary.recognized_revenue),
                _money(summary.recognized_production),
                _money(summary.total_commissions),
                _money(summary.net_billing),
            ],
        ],
        colWidths=[65 * mm] * 4,
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#44639F")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B9C7E2")),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 7 * mm), Paragraph("Composição", styles["Heading2"])])
    rows = [["Beneficiário", "Função/regra", "Calculada", "Adiado", "Pago", "A pagar", "Status"]]
    rows.extend(
        [
            item.beneficiary_name,
            ", ".join(item.strategies),
            _money(item.calculated_amount),
            _money(item.deferred_amount),
            _money(item.paid_amount),
            _money(item.payable_amount),
            item.status or "Não gerado",
        ]
        for item in beneficiaries
    )
    table = Table(
        rows,
        repeatRows=1,
        colWidths=[48 * mm, 55 * mm, 31 * mm, 31 * mm, 31 * mm, 31 * mm, 28 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C4B80")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 1), (5, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B9C7E2")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6FA")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return output.getvalue()


def _money(value: Decimal) -> str:
    number = float(value)
    formatted = f"{abs(number):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"{'-' if number < 0 else ''}R$ {formatted}"
