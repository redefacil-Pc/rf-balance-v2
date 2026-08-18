from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.modules.commissions.application.queries.financial_report import (
    FinancialReportBeneficiary,
    FinancialReportSummary,
)
from app.modules.reporting.application.financial_report_documents import (
    financial_report_pdf,
    financial_report_xlsx,
)


def _summary() -> FinancialReportSummary:
    return FinancialReportSummary(
        gross_revenue=Decimal("1000.00"),
        receipt_reversals=Decimal("100.00"),
        recognized_revenue=Decimal("900.00"),
        recognized_production=Decimal("5000.00"),
        consultant_commissions=Decimal("60.00"),
        leader_commissions=Decimal("10.00"),
        finalization_commissions=Decimal("20.00"),
        finalization_leader_commissions=Decimal("5.00"),
        bko_commissions=Decimal("15.00"),
        total_commissions=Decimal("110.00"),
        net_billing=Decimal("790.00"),
        bonuses=Decimal("0.00"),
        discounts=Decimal("0.00"),
        deferred=Decimal("0.00"),
        paid=Decimal("50.00"),
        payable=Decimal("60.00"),
    )


def _beneficiaries() -> list[FinancialReportBeneficiary]:
    return [
        FinancialReportBeneficiary(
            beneficiary_id=1,
            beneficiary_name="Carla Consultora",
            strategies=("STANDARD_CONSULTANT",),
            automatic_amount=Decimal("60.00"),
            manual_amount=Decimal("0.00"),
            calculated_amount=Decimal("60.00"),
            carryover_amount=Decimal("0.00"),
            bonus_amount=Decimal("0.00"),
            discount_amount=Decimal("0.00"),
            deferred_amount=Decimal("0.00"),
            paid_amount=Decimal("0.00"),
            payable_amount=Decimal("60.00"),
            status="PENDING",
        )
    ]


def test_xlsx_usa_o_resumo_e_a_composicao_do_mesmo_dto() -> None:
    content = financial_report_xlsx(_summary(), _beneficiaries())

    workbook = load_workbook(BytesIO(content), data_only=True)
    assert workbook.sheetnames == ["Resumo", "Beneficiários"]
    assert workbook["Resumo"]["B4"].value == 900
    assert workbook["Beneficiários"]["A2"].value == "Carla Consultora"
    assert workbook["Beneficiários"]["K2"].value == 60


def test_pdf_e_um_documento_valido_com_conteudo() -> None:
    content = financial_report_pdf(_summary(), _beneficiaries())

    assert content.startswith(b"%PDF-")
    assert len(content) > 1_000
